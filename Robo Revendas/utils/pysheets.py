from openpyxl import Workbook, load_workbook
from os import path

class Sheets:
    def __init__(self, output_folder, filename, fieldlist: list):
        self.output_folder = output_folder
        self.filename = filename
        if path.exists(f"{output_folder}/{filename}"):
            self.wb = load_workbook(f"{output_folder}/{filename}")
        else:
            self.wb = Workbook()
            self.ws = self.wb.active
            letters = list(map(chr, range(ord('A'), ord('Z') + 1)))
            index = 0
            for letter in letters[:len(fieldlist)]:
                self.ws[f'{letter}1'] = fieldlist[index]
                index += 1

        # self.save()

    def save_content(self, **kwargs):
        self.ws = self.wb.active
        values = []
        for key, value in kwargs.items():
            values.append(value)

        self.ws.append(values)

        self.save()

    def save(self):
        self.wb.save(filename=f"{self.output_folder}/{self.filename}")